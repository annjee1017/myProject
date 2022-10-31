# import the necessary packages
import numpy as np
import argparse
import cv2
import imutils
import os, time
from imutils.paths import list_images
from openpyxl import Workbook
import openpyxl
from openpyxl.drawing.image import Image

# hyper-parameters ---------------------------------
TEST_IMAGE_DIR = 'test/ok'
CAMNUM = 'test'
OUTPUT_DIR = 'test_output/test/ok'
BINARY_S, BINARY_E = 93, 255 # cam3 95 / cam3 110
#---------------------------------------------------

write_wb = Workbook()
write_ws = write_wb.active
write_ws['A1'] = 'NO'   # 1
write_ws['B1'] = 'PATH' # 2
write_ws['C1'] = '2'    # 3
write_ws['D1'] = '0'    # 4
write_ws['E1'] = '2'    # 5
write_ws['F1'] = '2'    # 6
write_ws['G1'] = '.'    # 7 
write_ws['H1'] = '1'    # 8
write_ws['I1'] = '0'    # 9
write_ws['J1'] = '.'    # 10
write_ws['K1'] = '1'    # 11
write_ws['L1'] = '0'    # 12
write_ws['M1'] = 'ㄲ'   # 13
write_ws['N1'] = 'ㅏ'   # 14
write_ws['O1'] = 'ㅈ'   # 15
write_ws['P1'] = 'ㅣ'   # 16

def imread(filename, dtype=np.uint8):
	try:
		n = np.fromfile(filename, dtype)
		img = cv2.imdecode(n, cv2.IMREAD_COLOR)
		return img
	except Exception as e:
		print(e)
		return None

def imwrite(filename, img, params=None):
	try:
		ext = os.path.splitext(filename)[1]
		result, n = cv2.imencode(ext, img, params)

		if result:
			with open(filename, mode='w+b') as f:
				n.tofile(f)
			return True
		else:
			return False
	except Exception as e:
		print(e)
		return False

line = 1
for num, path in enumerate(sorted(list(list_images(TEST_IMAGE_DIR)))):
    # time.sleep(0.02)
    foldername = path.split(os.sep)[-2]
    filename = path.split(os.sep)[-1]
    print(f"\n{foldername}/{filename}")

    write_ws.cell(line, 1, num+1) 
    write_ws.cell(line, 2, filename)

    # load image
    # cnt=0; cY_all=[]
    image = imread(path)
    output = image.copy()
    H, W = image.shape[:2]

    # image pre-processing
    thresh = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # while BINARY_S < 115:
    # ret, binary = cv2.threshold(thresh, -1, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
    ret, binary = cv2.threshold(thresh, BINARY_S, BINARY_E, cv2.THRESH_BINARY_INV)
    # binary = cv2.adaptiveThreshold(binary, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, \
    #                                  cv2.THRESH_BINARY, 3, 2)    
    
    crop = binary.copy()

    # find external contours in the image
    # cnts = cv2.findContours(binary, cv2.RETR_EXTERNAL , cv2.CHAIN_APPROX_SIMPLE)
    cnts = cv2.findContours(binary, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    contours = imutils.grab_contours(cnts)
    # sort contours
    sort_value = np.array([cv2.boundingRect(i)[0] for i in contours]).argsort()
    sort_contour = np.array(contours)[sort_value]

    count = 0
    line_in = line
    line_in+=1
    for (i, c) in enumerate(sort_contour):
        count +=1
        x, y, w, h = cv2.boundingRect(c)

        # check white pixels in each area (white 가 글자)
        crop = binary[y-2:y+h+2, x-2:x+w+2]
        black_count = np.sum(crop == 0)
        white_count = np.sum(crop == 255)
        print(f"{count}: {x} > black: {black_count}, white: {white_count}")

        if count == 1:
            write_ws.cell(line, 3, white_count)
        if count == 2:
            write_ws.cell(line, 4, white_count)
        if count == 3:
            write_ws.cell(line, 5, white_count)
        if count == 4:
            write_ws.cell(line, 6, white_count)
        if count == 5:    
            write_ws.cell(line, 7, white_count)
        if count == 6:    
            write_ws.cell(line, 8, white_count)
        if count == 7:            
            write_ws.cell(line, 9, white_count)
        if count == 8:                        
            write_ws.cell(line, 10, white_count)
        if count == 9:                        
            write_ws.cell(line, 11, white_count)
        if count == 10:                        
            write_ws.cell(line, 12, white_count)
        if count == 11:                        
            write_ws.cell(line, 13, white_count)
        if count == 12:                        
            write_ws.cell(line, 14, white_count)
        if count == 13:                        
            write_ws.cell(line, 15, white_count)
        if count == 14:                        
            write_ws.cell(line, 16, white_count)

        # draw image
        cv2.rectangle(output, (x-2,y-2), (x+w+2,y+h+2), (0,255,0), 2)
        cv2.putText(output, f"{count}", (x-2, 10), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
    
    # show output
    binary = imutils.resize(binary, width=1024)
    output = imutils.resize(output, width=1024)
    image = imutils.resize(image, width=1024)
    cv2.imshow("binary", binary)
    cv2.imshow("Contours", output)
    cv2.imshow("image", image)
    
    # save output
    os.makedirs(OUTPUT_DIR, exist_ok= True)
    os.makedirs(f'{OUTPUT_DIR}/check', exist_ok= True)
    line = line_in

    # save outliers
    if count != 14:
        key = cv2.waitKey(1)
        imwrite(f'{OUTPUT_DIR}/check/{filename[:-4]}_{BINARY_S}_0.jpg', binary)
        imwrite(f'{OUTPUT_DIR}/check/{filename[:-4]}_{BINARY_S}_1.jpg', output)
    else:
        key = cv2.waitKey(1)        
        imwrite(f'{OUTPUT_DIR}/{filename[:-4]}_{BINARY_S}_0.jpg', binary)
        imwrite(f'{OUTPUT_DIR}/{filename[:-4]}_{BINARY_S}_1.jpg', output)
    if key == 'q':
        break

write_wb.save(f'03_{CAMNUM}.xlsx')

